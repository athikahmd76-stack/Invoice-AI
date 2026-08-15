"""Excel / CSV export and Excel import."""

from __future__ import annotations

import csv
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any

from sqlalchemy import select
from sqlalchemy.orm import Session

from database import repository
from database.database import json_loads
from database.models import Invoice, InvoiceItem
from utils.formatting import parse_date, to_float, to_str

HEADER_FILL = "1F3864"
HEADER_FONT_COLOR = "FFFFFF"
CURRENCY_FMT = '#,##0.00;[Red]-#,##0.00'


def _style_worksheet(ws, headers: list[str], widths: dict[str, int]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    font = Font(color=HEADER_FONT_COLOR, bold=True, size=11)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = widths.get(header, 18)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _add_rows(ws, rows: list[dict], money_cols: set[str]) -> None:
    from openpyxl.styles import Alignment

    headers = list(rows[0].keys()) if rows else []
    for row in rows:
        values = []
        for key, value in row.items():
            if key in money_cols and isinstance(value, (int, float)):
                values.append(float(value))
            else:
                values.append(value)
        ws.append(values)
    for col_idx, key in enumerate(headers, start=1):
        column_letter = ws.cell(row=1, column=col_idx).column_letter
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if key in money_cols:
                cell.number_format = CURRENCY_FMT
                cell.alignment = Alignment(horizontal="right")


def export_excel(session: Session, filters: dict | None = None, output_dir: Path | None = None) -> Path:
    """Build a styled multi-sheet workbook: Invoices, Invoice Items, Vendors, Tax Summary, Processing Log."""
    from openpyxl import Workbook

    invoices = repository.invoice_rows_for_export(session, filters)
    item_dicts = repository.invoice_item_rows_for_export(session, filters)
    items = item_dicts or []
    vendors = [
        {
            "Vendor": v.name, "Legal Name": v.legal_name, "GSTIN": v.gstin, "PAN": v.pan,
            "Phone": v.phone, "Email": v.email, "Address": v.address,
        }
        for v in repository.list_vendors(session)
    ]
    tax_rows = repository.valid_tax_rows(session)
    logs = repository.logs_for_export(session)

    wb = Workbook()

    ws_inv = wb.active
    ws_inv.title = "Invoices"
    money_cols = {"Subtotal", "Discount", "Taxable Value", "CGST", "SGST", "IGST", "UTGST", "CESS", "Round Off", "Grand Total", "Amount Paid", "Balance Due"}
    invoice_rows = []
    for inv in invoices:
        invoice_rows.append({
            "Invoice No": inv["invoice_number"],
            "Invoice Date": inv["invoice_date"],
            "Vendor": inv["vendor_name"],
            "Vendor GSTIN": inv.get("vendor_gstin"),
            "Buyer": inv.get("buyer_name"),
            "PO Number": inv["po_number"],
            "Currency": inv.get("currency") or "INR",
            "Subtotal": inv.get("subtotal"),
            "Discount": inv.get("discount"),
            "Taxable Value": inv.get("taxable_value"),
            "CGST": inv.get("cgst"),
            "SGST": inv.get("sgst"),
            "IGST": inv.get("igst"),
            "UTGST": inv.get("utgst"),
            "CESS": inv.get("cess"),
            "Round Off": inv.get("round_off"),
            "Grand Total": inv["grand_total"],
            "Amount Paid": inv.get("amount_paid"),
            "Balance Due": inv.get("balance_due"),
            "Status": inv["status"],
            "Validation": inv["validation_status"],
            "Duplicate": inv["duplicate_status"],
            "Processed Date": inv.get("processed_at"),
        })
    invoice_headers = list(invoice_rows[0].keys()) if invoice_rows else []
    _style_worksheet(ws_inv, invoice_headers, {"Invoice No": 22, "Invoice Date": 13, "Vendor": 28, "Vendor GSTIN": 18, "PO Number": 16, "Processed Date": 21})
    _add_rows(ws_inv, invoice_rows, money_cols)

    ws_items = wb.create_sheet("Invoice Items")
    item_money = {"Unit Price", "MRP", "Discount Amount", "Taxable Value", "CGST Amt", "SGST Amt", "IGST Amt", "CESS Amt", "Line Total"}
    item_rows = []
    for it in items:
        item_rows.append({
            "Invoice No": it["invoice_number"], "Invoice Date": it["invoice_date"], "Vendor": it["vendor_name"],
            "Line No": it["line_no"], "Product Name": it["product_name"], "SKU": it["sku"], "HSN": it["hsn"],
            "UOM": it["uom"], "Quantity": it["quantity"], "Free Qty": it["free_quantity"], "Unit Price": it["unit_price"],
            "MRP": it["mrp"], "Discount %": it["discount_pct"], "Discount Amount": it["discount_amount"],
            "Taxable Value": it["taxable_value"], "GST %": it["gst_pct"], "CGST %": it["cgst_pct"],
            "SGST %": it["sgst_pct"], "IGST %": it["igst_pct"], "CGST Amt": it["cgst_amount"],
            "SGST Amt": it["sgst_amount"], "IGST Amt": it["igst_amount"], "CESS Amt": it["cess_amount"],
            "Line Total": it["line_total"],
        })
    item_headers = list(item_rows[0].keys()) if item_rows else ["Invoice No", "Product Name", "SKU", "HSN", "Quantity", "Unit Price", "Line Total"]
    _style_worksheet(ws_items, item_headers, {"Invoice No": 22, "Invoice Date": 13, "Vendor": 28, "Product Name": 34, "SKU": 18, "HSN": 14, "Line Total": 16, "Unit Price": 14})
    _add_rows(ws_items, item_rows, item_money)

    ws_vendors = wb.create_sheet("Vendors")
    vendor_headers = ["Vendor", "Legal Name", "GSTIN", "PAN", "Phone", "Email", "Address"]
    _style_worksheet(ws_vendors, vendor_headers, {"Vendor": 28, "Legal Name": 28, "GSTIN": 18, "PAN": 14, "Phone": 16, "Email": 26, "Address": 40})
    _add_rows(ws_vendors, vendors, set())

    ws_tax = wb.create_sheet("Tax Summary")
    tax_headers = ["Invoice No", "Tax Type", "Rate %", "Taxable Value", "Amount", "Source"]
    tax_out = []
    for t in tax_rows:
        invoice = session.get(Invoice, t["invoice_id"])
        tax_out.append({
            "Invoice No": invoice.invoice_number if invoice else None, "Tax Type": t["tax_type"],
            "Rate %": t["rate"], "Taxable Value": t["taxable_value"], "Amount": t["amount"], "Source": t["source"],
        })
    _style_worksheet(ws_tax, tax_headers, {"Invoice No": 22, "Tax Type": 12, "Amount": 16})
    _add_rows(ws_tax, tax_out, {"Taxable Value", "Amount"})

    ws_logs = wb.create_sheet("Processing Log")
    log_headers = ["File Name", "Upload Date", "Start Time", "End Time", "Duration (ms)", "OCR Status", "AI Status", "Validation", "Model Used", "Status", "Error", "Retries"]
    log_rows = [
        {
            "File Name": l["file_name"], "Upload Date": l["upload_date"], "Start Time": l["start_time"],
            "End Time": l["end_time"], "Duration (ms)": l["duration_ms"], "OCR Status": l["ocr_status"],
            "AI Status": l["ai_status"], "Validation": l["validation_status"], "Model Used": l["model_used"],
            "Status": l["status"], "Error": l["error_message"], "Retries": l["retry_count"],
        }
        for l in logs
    ]
    _style_worksheet(ws_logs, log_headers, {"File Name": 34, "Error": 50, "Model Used": 18})
    _add_rows(ws_logs, log_rows, {"Duration (ms)"})

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"invoiceai_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(out_path)
    return out_path


def export_csv(session: Session, filters: dict | None = None, output_dir: Path | None = None) -> dict[str, Path]:
    """Export Invoices and Line Items as CSV files."""
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    paths = {}
    invoices = repository.invoice_rows_for_export(session, filters)
    items = repository.invoice_item_rows_for_export(session, filters)

    inv_path = output_dir / f"invoices_{stamp}.csv"
    with open(inv_path, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)
        writer.writerow(["Invoice No", "Invoice Date", "Vendor", "Vendor GSTIN", "PO Number", "Status", "Validation", "Duplicate", "Grand Total"])
        for inv in invoices:
            writer.writerow([
                inv["invoice_number"], inv["invoice_date"], inv["vendor_name"], inv.get("vendor_gstin"),
                inv["po_number"], inv["status"], inv["validation_status"], inv["duplicate_status"],
                inv["grand_total"],
            ])
    paths["invoices"] = inv_path

    item_path = output_dir / f"invoice_items_{stamp}.csv"
    with open(item_path, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)
        writer.writerow(["Invoice No", "Invoice Date", "Vendor", "Line No", "Product Name", "SKU", "HSN", "Qty", "UOM", "Unit Price", "GST %", "Taxable Value", "Line Total"])
        for it in items:
            writer.writerow([
                it["invoice_number"], it["invoice_date"], it["vendor_name"], it["line_no"],
                it["product_name"], it["sku"], it["hsn"], it["quantity"], it["uom"],
                it["unit_price"], it["gst_pct"], it["taxable_value"], it["line_total"],
            ])
    paths["items"] = item_path
    return paths


def excel_bytes(session: Session, filters: dict | None = None) -> bytes:
    """Workbook bytes for st.download_button."""
    with tempfile.TemporaryDirectory() as tmp:
        path = export_excel(session, filters, Path(tmp))
        return path.read_bytes()


# ==================================================== complete report (items)

def calc_cost_price(vendor_name: str | None, item) -> float | None:
    """Vendor-specific cost price logic (unchanged from the legacy tool)."""
    vendor = (vendor_name or "").strip().lower()
    qty = item.quantity
    if qty is None or qty <= 0:
        return None
    if "extreme adventure sports" in vendor:
        taxable = item.taxable_value
        if taxable is None:
            return None
        return round(taxable / qty, 2)
    if "gillidanda" in vendor:
        amount = item.line_total
        tax_rate = item.gst_pct
        if amount is None or not tax_rate:
            return None
        return round(amount / qty / (1 + float(tax_rate) / 100), 2)
    return None


def complete_report_excel(session: Session, output_dir: Path | None = None) -> Path:
    """Single-sheet workbook: every line item across all invoices."""
    from openpyxl import Workbook

    rows = session.execute(
        select(InvoiceItem, Invoice)
        .join(Invoice, InvoiceItem.invoice_id == Invoice.id)
        .order_by(InvoiceItem.invoice_id, InvoiceItem.line_no)
    ).all()
    data = []
    for item, inv in rows:
        bill_to = inv.buyer_name or ""
        if inv.buyer_address:
            bill_to = f"{bill_to}, {inv.buyer_address}" if bill_to else inv.buyer_address
        vendor_name = inv.vendor_name
        if not vendor_name and inv.raw_json:
            vendor_name = (json_loads(inv.raw_json).get("vendor") or {}).get("name")
        data.append({
            "Invoice Number": inv.invoice_number,
            "Invoice Date": inv.invoice_date,
            "Vendor Name": vendor_name,
            "PO Number": inv.po_number,
            "Bill To": bill_to,
            "File Name": inv.original_filename,
            "Total Amount": inv.grand_total,
            "SKU": item.sku,
            "Product Code": item.product_code,
            "Product Name": item.product_name,
            "EAN": item.ean,
            "HSN": item.hsn,
            "Cost Price": calc_cost_price(inv.vendor_name, item),
            "Quantity": item.quantity,
            "Unit Price": item.unit_price,
            "Taxable Value": item.taxable_value,
            "Tax Rate %": item.gst_pct,
            "Line Total": item.line_total,
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "Complete Report"
    headers = [
        "Invoice Number", "Invoice Date", "Vendor Name", "PO Number", "Bill To",
        "File Name", "Total Amount", "SKU", "Product Code", "Product Name", "EAN", "HSN", "Cost Price",
        "Quantity", "Unit Price", "Taxable Value", "Tax Rate %", "Line Total",
    ]
    _style_worksheet(ws, headers, {
        "Invoice Number": 20, "Invoice Date": 13, "Vendor Name": 30, "PO Number": 16, "Bill To": 50,
        "File Name": 24, "Total Amount": 14, "SKU": 20, "Product Code": 20,
        "Product Name": 44, "EAN": 16, "HSN": 12, "Cost Price": 13, "Quantity": 11, "Unit Price": 13,
        "Taxable Value": 14, "Tax Rate %": 11, "Line Total": 14,
    })
    _add_rows(ws, data, {"Total Amount", "Cost Price", "Unit Price", "Taxable Value", "Line Total"})

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"invoiceai_complete_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(out_path)
    return out_path


def complete_report_bytes(session: Session) -> bytes:
    """Complete-report workbook bytes for st.download_button."""
    with tempfile.TemporaryDirectory() as tmp:
        path = complete_report_excel(session, Path(tmp))
        return path.read_bytes()


# ==================================================================== import

def _map_header(header: Any) -> str:
    h = str(header).strip().lower().replace(" ", "_").replace("/", "_").replace("-", "_")
    mapping = {
        "invoice_no": "invoice_number", "invoice_number": "invoice_number", "inv_no": "invoice_number",
        "invoice": "invoice_number", "inv_number": "invoice_number",
        "invoice_date": "invoice_date", "inv_date": "invoice_date", "date": "invoice_date",
        "vendor": "vendor_name", "vendor_name": "vendor_name", "supplier": "vendor_name", "party_name": "vendor_name",
        "vendor_gstin": "vendor_gstin", "gstin": "vendor_gstin", "gst_no": "vendor_gstin",
        "po_number": "po_number", "po_no": "po_number", "purchase_order": "po_number",
        "total": "grand_total", "grand_total": "grand_total", "total_amount": "grand_total",
        "invoice_amount": "grand_total", "amount": "grand_total", "invoice_value": "grand_total",
        "tax": "tax", "total_tax": "tax", "tax_amount": "tax",
        "items": "item_count", "no_of_items": "item_count",
    }
    return mapping.get(h, h)


def import_invoices_excel(session: Session, path: Path) -> dict:
    """Import an invoice master file (xlsx/xls) into local records with status 'Imported'."""
    import pandas as pd

    from services.duplicate_service import compute_fingerprint

    try:
        df = pd.read_excel(path)
    except Exception as exc:
        raise ValueError(f"Could not read Excel file: {exc}") from exc
    if df.empty:
        return {"imported": 0, "skipped": 0, "errors": []}

    df.columns = [_map_header(c) for c in df.columns]
    imported = 0
    skipped = 0
    errors: list[str] = []
    for idx, row in df.iterrows():
        data = {col: (None if pd.isna(val) else val) for col, val in row.items()}
        invoice_number = to_str(data.get("invoice_number"))
        if not invoice_number:
            skipped += 1
            continue
        vendor_name = to_str(data.get("vendor_name"))
        vendor_gstin = to_str(data.get("vendor_gstin"))
        try:
            parsed = {
                "vendor": {"name": vendor_name, "gstin": vendor_gstin},
                "invoice": {
                    "number": invoice_number,
                    "date": parse_date(data.get("invoice_date")),
                    "po_number": to_str(data.get("po_number")),
                },
                "buyer": {},
                "items": [],
                "taxes": {},
                "totals": {"grand_total": to_float(data.get("grand_total"))},
                "payment": {},
                "confidence": {},
            }
            fingerprint = compute_fingerprint(vendor_gstin, invoice_number)
            existing = None
            if fingerprint:
                existing = session.execute(
                    select(Invoice).where(Invoice.fingerprint == fingerprint)
                ).scalars().first()
            duplicate = {"status": "None", "fingerprint": fingerprint, "duplicate_of_id": None}
            if existing:
                duplicate["status"] = "Duplicate"
                duplicate["duplicate_of_id"] = existing.id
                tax_value = to_float(data.get("tax"))
                if tax_value is not None:
                    parsed["taxes"]["igst"] = tax_value
            repository.create_invoice(
                session,
                parsed,
                {"status": "Imported", "validation_status": None, "checks": []},
                duplicate,
                {
                    "original_filename": Path(path).name,
                    "file_path": None,
                    "file_hash": None,
                    "page_count": None,
                    "preview_path": None,
                    "ai_model": "excel-import",
                    "ai_mode": "text",
                    "ocr_engine": "excel-import",
                },
            )
            imported += 1
        except Exception as exc:  # noqa: BLE001
            errors.append(f"Row {idx + 2}: {exc}")
            skipped += 1
    session.flush()
    return {"imported": imported, "skipped": skipped, "errors": errors}